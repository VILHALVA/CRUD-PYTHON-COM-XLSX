import os
from openpyxl import Workbook, load_workbook
from time import sleep

class Usuario:
    def __init__(self, nome, idade='', email='', telefone=''):
        self.nome = nome
        self.idade = idade
        self.email = email
        self.telefone = telefone

class GerenciadorUsuarios:
    def __init__(self):
        self.arquivo = os.path.join(os.path.dirname(__file__), 'usuarios.xlsx')
        if not os.path.exists(self.arquivo):
            wb = Workbook()
            ws = wb.active
            ws.title = "Usuarios"
            ws.append(["Nome", "Idade", "Email", "Telefone"])
            wb.save(self.arquivo)

    def adicionar_usuario(self, nome, idade='', email='', telefone=''):
        wb = load_workbook(self.arquivo)
        ws = wb["Usuarios"]
        ws.append([nome, idade, email, telefone])
        wb.save(self.arquivo)
        print("😎USUÁRIO ADICIONADO COM SUCESSO!")

    def listar_usuarios(self):
        if os.path.exists(self.arquivo):
            wb = load_workbook(self.arquivo)
            ws = wb["Usuarios"]
            print("=" * 100)
            print("LISTA DE USUÁRIOS:")
            print("-" * 100)
            for row in ws.iter_rows(min_row=2, values_only=True):
                print("*" * 100)
                print(f"NOME: {row[0]}, IDADE: {row[1]}, EMAIL: {row[2]}, TELEFONE: {row[3]}")
                print("*" * 100)
            print("=" * 100)
        else:
            print("😒NENHUM USUÁRIO CADASTRADO.")

    def atualizar_usuario(self, nome_antigo, novo_nome, nova_idade='', novo_email='', novo_telefone=''):
        wb = load_workbook(self.arquivo)
        ws = wb["Usuarios"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == nome_antigo:
                row[0].value = novo_nome
                if nova_idade:
                    row[1].value = nova_idade
                if novo_email:
                    row[2].value = novo_email
                if novo_telefone:
                    row[3].value = novo_telefone
                break
        wb.save(self.arquivo)
        print("😙USUÁRIO ATUALIZADO COM SUCESSO!")

    def excluir_usuario(self, nome):
        wb = load_workbook(self.arquivo)
        ws = wb["Usuarios"]
        rows = list(ws.iter_rows(min_row=2))
        for row in rows:
            if row[0].value == nome:
                ws.delete_rows(row[0].row)
                break
        wb.save(self.arquivo)
        print("🗑USUÁRIO EXCLUÍDO COM SUCESSO!")

def exibir_menu():
    print("\nMENU:")
    print("1. ADICIONAR USUÁRIO")
    print("2. LISTAR USUÁRIOS")
    print("3. ATUALIZAR USUÁRIO")
    print("4. EXCLUIR USUÁRIO")
    print("5. SAIR")

def main():
    gerenciador = GerenciadorUsuarios()

    while True:
        exibir_menu()
        opcao = input("😎ESCOLHA UMA OPÇÃO:\n>>> ")

        if opcao == "1":
            nome = input("😎DIGITE O NOME:\n>>> ")
            idade = input("😎DIGITE A IDADE (Pressione Enter para deixar em branco):\n>>> ")
            email = input("😎DIGITE O EMAIL (Pressione Enter para deixar em branco):\n>>> ")
            telefone = input("😎DIGITE O TELEFONE (Pressione Enter para deixar em branco):\n>>> ")
            gerenciador.adicionar_usuario(nome, idade, email, telefone)
        elif opcao == "2":
            gerenciador.listar_usuarios()
        elif opcao == "3":
            nome_antigo = input("😎DIGITE O NOME A SER ATUALIZADO:\n>>> ")
            novo_nome = input("😎DIGITE O NOVO NOME:\n>>> ")
            nova_idade = input("😎DIGITE A NOVA IDADE (Pressione Enter para deixar em branco):\n>>> ")
            novo_email = input("😎DIGITE O NOVO EMAIL (Pressione Enter para deixar em branco):\n>>> ")
            novo_telefone = input("😎DIGITE O NOVO TELEFONE (Pressione Enter para deixar em branco):\n>>> ")
            gerenciador.atualizar_usuario(nome_antigo, novo_nome, nova_idade, novo_email, novo_telefone)
        elif opcao == "4":
            nome = input("😎DIGITE O NOME DO USUÁRIO A SER EXCLUÍDO:\n>>> ")
            gerenciador.excluir_usuario(nome)
        elif opcao == "5":
            print("🚀SAINDO...")
            sleep(3)
            break
        else:
            print("😡OPÇÃO INVÁLIDA. TENTE NOVAMENTE!")

if __name__ == "__main__":
    main()
